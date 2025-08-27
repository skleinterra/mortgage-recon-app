import io, re, os
import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook
from rapidfuzz import fuzz, process

EXPECTED_HEADERS=[
 "Property","Mortgage 1st","Mortgage 2nd","Interest Mortgage 1st","Interest Mortgage 2nd",
 "Tax Escrow","Escrow-Insurance","Escrow-Interest Reserve","Escrow-Debt Service Reserve",
 "Escrow-Immediate Replacement Reserve","Escrow-Replacement Reserve","Escrow-Renovation Reserve","Other Escrows"
]
HEADER_SYNONYMS={
 "insurance escrow":"Escrow-Insurance","reserves bal":"Other Escrows","reserves balance":"Other Escrows",
 "reserve balance":"Other Escrows","tax escrow":"Tax Escrow","principal balance 1st":"Mortgage 1st",
 "principal balance 2nd":"Mortgage 2nd","interest 1st":"Interest Mortgage 1st","interest 2nd":"Interest Mortgage 2nd"
}
LINEVALS=[re.compile(r"(?P<label>[A-Za-z \-_/&]+)[:\s]+(?P<val>\(?\$?[\d,]+(?:\.\d{1,2})?\)?)",re.I)]

def _norm(s): return re.sub(r"\s+"," ",(s or "").strip().lower().replace("_"," ").replace("-"," "))
def _parse_num(s):
    if not s: return None
    s=s.strip(); neg=s.startswith("(") and s.endswith(")")
    if neg: s=s[1:-1]
    s=s.replace("$","").replace(",","")
    try: v=float(s); return -v if neg else v
    except: return None

def _label_val(line:str):
    for rx in LINEVALS:
        m=rx.search(line)
        if m:
            v=_parse_num(m.group("val"))
            if v is not None: return m.group("label").strip(), v
    toks=line.rsplit(" ",1)
    if len(toks)==2:
        v=_parse_num(toks[1]); 
        if v is not None: return toks[0].strip(" :.-"), v
    return None

def _detect_vendor(full_text:str, vendor_df:pd.DataFrame):
    if vendor_df is None or vendor_df.empty: return None
    scores={}
    for v,grp in vendor_df.groupby("Vendor"):
        score=0
        if "DetectPattern" in grp.columns:
            for pat in grp["DetectPattern"]:
                if pat and re.search(pat,full_text,re.I|re.M): score+=5
        for pat in grp["Pattern"]:
            if pat and str(pat).lower() in full_text.lower(): score+=1
        if score: scores[v]=score
    return max(scores,key=scores.get) if scores else None

def _map_header(lbl:str, vendor:str|None, vendor_df:pd.DataFrame):
    if vendor_df is not None and not vendor_df.empty:
        sub=vendor_df[vendor_df["Vendor"]==vendor] if vendor else vendor_df
        for _,r in sub.iterrows():
            pat=str(r.get("Pattern","")); hdr=str(r.get("MappedHeader",""))
            if pat and pat.lower() in lbl.lower() and hdr in EXPECTED_HEADERS: return hdr
    n=_norm(lbl)
    if n in HEADER_SYNONYMS: return HEADER_SYNONYMS[n]
    for h in EXPECTED_HEADERS:
        if _norm(h)==n: return h
    if "reserve" in n: return "Other Escrows"
    return None

def _resolve_property(full_text:str, name2code:dict, code2name:dict):
    for c in code2name:
        if re.search(rf"\b{re.escape(c)}\b",full_text,re.I): return c
    for n,c in name2code.items():
        if re.search(rf"\b{re.escape(n)}\b",full_text,re.I): return c
    m=process.extractOne(full_text[:5000],list(name2code),scorer=fuzz.partial_ratio)
    if m and m[1] >= 92: return name2code[m[0]]
    raise RuntimeError("Clarification needed: PropertyCode not found")

def _read_text_with_plumber(pdf_bytes:bytes):
    import warnings, io as _io
    text_lines=[]; full=[]
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        with pdfplumber.open(_io.BytesIO(pdf_bytes)) as doc:
            for p in doc.pages:
                t=p.extract_text() or ""
                if t:
                    full.append(t)
                    text_lines+= [ln.strip() for ln in t.splitlines() if ln.strip()]
    return "\n".join(full), text_lines

# ---- OCR Adapters (choose via env OCR_PROVIDER = 'gcv' or 'azure') ----
def _ocr_google(pdf_bytes:bytes) -> tuple[str,list[str]]:
    # GOOGLE_APPLICATION_CREDENTIALS_FILE is created in app at runtime from secret JSON
    from google.cloud import vision
    client = vision.ImageAnnotatorClient()
    image = vision.Image(content=pdf_bytes)
    resp = client.document_text_detection(image=image)
    if resp.error.message: raise RuntimeError(resp.error.message)
    text = (resp.full_text_annotation.text or "").strip()
    return text, [ln.strip() for ln in text.splitlines() if ln.strip()]

def _ocr_azure(pdf_bytes:bytes) -> tuple[str,list[str]]:
    from azure.ai.documentintelligence import DocumentIntelligenceClient
    from azure.core.credentials import AzureKeyCredential
    endpoint = os.environ["AZURE_DOCUMENTINTELLIGENCE_ENDPOINT"]
    key = os.environ["AZURE_DOCUMENTINTELLIGENCE_KEY"]
    client = DocumentIntelligenceClient(endpoint=endpoint, credential=AzureKeyCredential(key))
    poller = client.begin_analyze_document("prebuilt-read", pdf_bytes)
    result = poller.result()
    lines=[]
    for page in getattr(result,"pages",[]):
        for ln in getattr(page,"lines",[]):
            if ln.content: lines.append(ln.content.strip())
    return "\n".join(lines), lines

def extract_text_or_ocr(pdf_bytes:bytes) -> tuple[str,list[str]]:
    full, lines = _read_text_with_plumber(pdf_bytes)
    if full.strip() and len(lines) >= 3: return full, lines
    provider = os.environ.get("OCR_PROVIDER","gcv").lower()
    if provider == "azure": return _ocr_azure(pdf_bytes)
    return _ocr_google(pdf_bytes)

def _prep_template(xlsx_bytes:bytes|None):
    if xlsx_bytes:
        wb=load_workbook(io.BytesIO(xlsx_bytes))
        if "Mortgage Import" not in wb.sheetnames: wb.create_sheet("Mortgage Import",0)
    else:
        wb=Workbook(); wb.create_sheet("Mortgage Import",0)
    ws=wb["Mortgage Import"]
    for row in ws.iter_rows(min_row=5, max_row=max(ws.max_row,5)):
        for c in row: c.value=None
    ws.cell(4,1,"#")
    for i,h in enumerate(EXPECTED_HEADERS, start=2): ws.cell(4,i,h)
    return wb, ws

def run_pipeline_in_memory(pdf_blobs:list[tuple[str,bytes]],
                           datagrid_df:pd.DataFrame,
                           vendor_df:pd.DataFrame|None,
                           template_bytes:bytes|None) -> io.BytesIO:
    if not {"PropertyCode","PropertyName"}.issubset(datagrid_df.columns):
        raise RuntimeError("DataGridExport.csv must include PropertyCode,PropertyName")
    name2code=dict(zip(datagrid_df["PropertyName"].astype(str), datagrid_df["PropertyCode"].astype(str)))
    code2name=dict(zip(datagrid_df["PropertyCode"].astype(str), datagrid_df["PropertyName"].astype(str)))

    wb, ws = _prep_template(template_bytes)
    for row in ws.iter_rows(min_row=5, max_row=max(ws.max_row,5)):
        for c in row: c.value=None

    r, cnt = 5, 1
    for fname, blob in pdf_blobs:
        full, lines = extract_text_or_ocr(blob)
        vendor = _detect_vendor(full, vendor_df) if vendor_df is not None else None
        prop = _resolve_property(full, name2code, code2name)  # raises single clarification if needed
        for ln in lines:
            lv = _label_val(ln)
            if not lv: continue
            hdr = _map_header(lv[0], vendor, vendor_df)
            if not hdr: continue
            col = EXPECTED_HEADERS.index(hdr)+2
            ws.cell(r,1,cnt); ws.cell(r,2,prop); ws.cell(r,col,lv[1])
            r += 1; cnt += 1

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out
